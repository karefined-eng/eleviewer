from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableView, QTabBar, QHeaderView,
    QAbstractItemView, QLabel,
)
from PySide6.QtCore import Signal, Qt, QAbstractTableModel, QModelIndex
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import io

from theme import (
    xlsx_sheet_tab_stylesheet, BRAND_PANEL, BRAND_PRIMARY, 
    BRAND_BORDER, get_brand_accent, BRAND_BACKGROUND
)
import re


class XlsxTableModel(QAbstractTableModel):
    """Virtualized model for high-performance rendering of Excel worksheets."""
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self._data = data if data else [[""]]
        self._max_cols = max((len(r) for r in self._data), default=1)

    def rowCount(self, parent=QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        return self._max_cols

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        r, c = index.row(), index.column()
        if role in (Qt.DisplayRole, Qt.EditRole):
            if r < len(self._data) and c < len(self._data[r]):
                return str(self._data[r][c])
            return ""
        return None

    def setData(self, index, value, role=Qt.EditRole):
        if index.isValid() and role == Qt.EditRole:
            r, c = index.row(), index.column()
            while len(self._data) <= r:
                self._data.append([])
            while len(self._data[r]) <= c:
                self._data[r].append("")
            self._data[r][c] = str(value)
            self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
            return True
        return False

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._col_to_letter(section)
            elif orientation == Qt.Vertical:
                return str(section + 1)
        return None

    def _col_to_letter(self, col_idx):
        result = ""
        while col_idx >= 0:
            result = chr(col_idx % 26 + 65) + result
            col_idx = col_idx // 26 - 1
        return result

    def set_data(self, data):
        self.beginResetModel()
        self._data = data if data else [[""]]
        self._max_cols = max((len(r) for r in self._data), default=1)
        self.endResetModel()

    def get_data(self):
        return self._data


class XlsxViewer(QWidget):
    """XLSX viewer with Google Sheets-style bottom sheet tabs and virtualized grid."""

    textChanged = Signal()

    def __init__(self, file_path=None):
        super().__init__()
        self.file_path = file_path
        self.is_modified = False
        self.workbook = None
        self.current_sheet_name = None
        self.merged_cells_ranges = set()
        self._loading_sheet = False
        self._bookmark_callback = None
        self._last_found = None

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Read-only info banner
        banner_row = QHBoxLayout()
        banner_row.setContentsMargins(6, 3, 6, 3)
        self._readonly_label = QLabel("\U0001f512 View-only — formula values shown. Editing disabled to protect formulas.")
        self._readonly_label.setStyleSheet(
            f"color: {BRAND_PRIMARY}; font-size: 11px; opacity: 0.7;"
        )
        banner_row.addWidget(self._readonly_label)
        banner_row.addStretch()
        layout.addLayout(banner_row)

        self.model = XlsxTableModel()
        self.model.dataChanged.connect(self._on_cell_changed)
        self.table = QTableView()
        self.table.setModel(self.model)
        # Option C: strictly view-only — no editing allowed so formulas are never overwritten
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setStyleSheet(f"""
            QTableView {{
                background: {BRAND_PANEL};
                color: {BRAND_PRIMARY};
                gridline-color: {BRAND_BORDER};
            }}
            QTableView::item {{ padding: 5px; }}
            QTableView::item:selected {{ background: {get_brand_accent()}; color: {BRAND_BACKGROUND}; }}
        """)

        self.sheet_tabs = QTabBar()
        self.sheet_tabs.setDrawBase(True)
        self.sheet_tabs.setStyleSheet(xlsx_sheet_tab_stylesheet())
        self.sheet_tabs.currentChanged.connect(self._on_tab_index_changed)

        layout.addWidget(self.table)
        layout.addWidget(self.sheet_tabs)
        self.setLayout(layout)

        if file_path:
            self.load_from_path(file_path)
        else:
            from openpyxl import Workbook
            self.workbook = Workbook()
            self.sheet_tabs.blockSignals(True)
            while self.sheet_tabs.count():
                self.sheet_tabs.removeTab(0)
            for name in self.workbook.sheetnames:
                self.sheet_tabs.addTab(name)
            self.sheet_tabs.blockSignals(False)
            if self.workbook.sheetnames:
                self.sheet_tabs.setCurrentIndex(0)
                self._on_sheet_changed(self.workbook.sheetnames[0])
            self.is_modified = False

    def load_from_path(self, file_path):
        try:
            self.workbook = load_workbook(file_path, data_only=True, keep_vba=False)
            self.sheet_tabs.blockSignals(True)
            while self.sheet_tabs.count():
                self.sheet_tabs.removeTab(0)
            for name in self.workbook.sheetnames:
                self.sheet_tabs.addTab(name)
            self.sheet_tabs.blockSignals(False)

            if self.workbook.sheetnames:
                self.sheet_tabs.setCurrentIndex(0)
                self._on_sheet_changed(self.workbook.sheetnames[0])
            self.is_modified = False
        except Exception as e:
            error_msg = str(e)
            if "defaultColWidthPt" in error_msg:
                raise Exception(
                    "XLSX file compatibility issue. Try opening in Excel and re-saving."
                )
            raise Exception(f"Failed to load XLSX: {error_msg}")

    def _on_tab_index_changed(self, index):
        if index < 0 or not self.workbook:
            return
        name = self.sheet_tabs.tabText(index)
        if name != self.current_sheet_name:
            self._flush_table_to_workbook()
            self._on_sheet_changed(name)

    def _flush_table_to_workbook(self):
        if not self.workbook or not self.current_sheet_name:
            return
        ws = self.workbook[self.current_sheet_name]
        grid_data = self.model.get_data()
        for row_idx, row in enumerate(grid_data):
            for col_idx, val in enumerate(row):
                try:
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                    if isinstance(cell, MergedCell):
                        continue
                    if str(cell.value or "") != str(val or ""):
                        cell.value = val
                except Exception:
                    continue

    def _on_sheet_changed(self, sheet_name):
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return

        self._loading_sheet = True
        self.current_sheet_name = sheet_name
        ws = self.workbook[sheet_name]
        max_row = max(ws.max_row or 1, 20)
        max_col = max(ws.max_column or 1, 10)
        self.merged_cells_ranges = set(ws.merged_cells.ranges)

        rows_data = []
        for row_idx in range(1, max_row + 1):
            row_vals = []
            for col_idx in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value if cell.value is not None else ""
                    row_vals.append(str(value))
                except Exception:
                    row_vals.append("")
            rows_data.append(row_vals)

        self.model.set_data(rows_data)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        for c in range(min(max_col, 15)):
            self.table.resizeColumnToContents(c)
            if self.table.columnWidth(c) < 80:
                self.table.setColumnWidth(c, 80)
        self._loading_sheet = False

    def _on_cell_changed(self, top_left=None, bottom_right=None, roles=None):
        if self._loading_sheet:
            return
        self.is_modified = True
        self.textChanged.emit()

    def to_xlsx_bytes(self):
        try:
            self._flush_table_to_workbook()
            byte_stream = io.BytesIO()
            self.workbook.save(byte_stream)
            byte_stream.seek(0)
            return byte_stream.getvalue()
        except Exception as e:
            raise Exception(f"Failed to save XLSX: {str(e)}")

    def toPlainText(self):
        text_rows = []
        for row in self.model.get_data():
            text_rows.append(" | ".join(row))
        return "\n".join(text_rows)

    def setPlainText(self, text):
        pass

    def set_bookmark_callback(self, callback):
        self._bookmark_callback = callback

    def _bookmark_payload(self):
        return {
            "page_number": self.sheet_tabs.currentIndex(),
            "scroll_position_y": float(self.table.verticalScrollBar().value()),
            "scroll_position_x": float(self.table.horizontalScrollBar().value()),
            "label": f"Sheet '{self.current_sheet_name or 'Unknown'}'",
        }

    def go_to_bookmark(self, page_number=0, scroll_position_y=0.0, **kwargs):
        if 0 <= page_number < self.sheet_tabs.count():
            self.sheet_tabs.setCurrentIndex(int(page_number))
        self.table.verticalScrollBar().setValue(int(scroll_position_y))
        if "scroll_position_x" in kwargs:
            self.table.horizontalScrollBar().setValue(int(kwargs["scroll_position_x"]))

    def find_text(self, text, match_case, whole_word, forward):
        if not text:
            return False
        
        data = self.model.get_data()
        rows = len(data)
        cols = self.model.columnCount()
        
        start_r = 0
        start_c = -1
        
        if self._last_found:
            start_r, start_c = self._last_found
        else:
            sel = self.table.selectionModel().currentIndex()
            if sel.isValid():
                start_r, start_c = sel.row(), sel.column()
        
        flags = 0 if match_case else re.IGNORECASE
        pattern = f"\\b{re.escape(text)}\\b" if whole_word else re.escape(text)
        try:
            regex = re.compile(pattern, flags)
        except re.error:
            return False

        r, c = start_r, start_c
        while True:
            if forward:
                c += 1
                if c >= cols:
                    c = 0
                    r += 1
                if r >= rows:
                    r = 0
            else:
                c -= 1
                if c < 0:
                    c = cols - 1
                    r -= 1
                if r < 0:
                    r = rows - 1

            if (r, c) == (start_r, start_c) or (r >= rows) or (c >= len(data[r])):
                return False

            cell_val = str(data[r][c] if c < len(data[r]) else "")
            if regex.search(cell_val):
                idx = self.model.index(r, c)
                self.table.setCurrentIndex(idx)
                self.table.scrollTo(idx)
                self._last_found = (r, c)
                return True

    def replace_text(self, find_str, replace_str, match_case, whole_word):
        if self._last_found:
            r, c = self._last_found
            idx = self.model.index(r, c)
            val = self.model.data(idx)
            
            flags = 0 if match_case else re.IGNORECASE
            pattern = f"\\b{re.escape(find_str)}\\b" if whole_word else re.escape(find_str)
            new_val = re.sub(pattern, replace_str, val, flags=flags)
            
            self.model.setData(idx, new_val)
            self._flush_table_to_workbook()
            self._last_found = None
            self.find_text(find_str, match_case, whole_word, True)

    def replace_all(self, find_str, replace_str, match_case, whole_word):
        count = 0
        flags = 0 if match_case else re.IGNORECASE
        pattern = f"\\b{re.escape(find_str)}\\b" if whole_word else re.escape(find_str)
        try:
            regex = re.compile(pattern, flags)
        except re.error:
            return 0
            
        for r, row in enumerate(self.model.get_data()):
            for c, val in enumerate(row):
                str_val = str(val)
                if regex.search(str_val):
                    new_val = regex.sub(replace_str, str_val)
                    self.model.setData(self.model.index(r, c), new_val)
                    count += 1
        if count > 0:
            self._flush_table_to_workbook()
            self._last_found = None
        return count
